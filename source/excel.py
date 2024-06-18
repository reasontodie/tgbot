from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def create_excel_file(filename) -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(["datetime", "vacancy_count", "change"])

    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for col in range(1, 4):
        ws.cell(row=1, column=col).fill = yellow_fill
    wb.save(filename)


def add_data_to_excel(filename, col1_data, col2_data, col3_data) -> None:
    try:
        wb = load_workbook(filename)
        ws = wb.active
    except FileNotFoundError:
        create_excel_file(filename)
        wb = load_workbook(filename)
        ws = wb.active

    ws.append([col1_data, col2_data, col3_data])
    wb.save(filename)
